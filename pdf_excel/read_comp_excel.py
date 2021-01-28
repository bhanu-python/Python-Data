import openpyxl

#load workbook

wb=openpyxl.load_workbook('data.xlsx')

#create worksheet object

ws=wb['Name']

#maximum num of row
row=ws.max_row

#maximum num of column
col=ws.max_column



for i in range(1,row+1):
    for j in range(1,col+1):
        print(ws.cell(i,j).value,end=' ')
    print('\n')
