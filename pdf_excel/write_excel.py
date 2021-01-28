import openpyxl

from openpyxl import Workbook

#creating a Excel work book

wb=Workbook()

#getting a active worksheet

ws=wb.active

#creating work sheet in excel workbook

ws1=wb.create_sheet('sheet 1')
ws2=wb.create_sheet('sheet 2')

#storing the data in column

#methode1 

ws['A1']=1
ws2['A1']=2

#methode2

ws.cell(2,2,value="bhanu")

var="bhanu"
ws.cell(2,3,value=var)

#saving the file

wb.save('test.xlsx')
