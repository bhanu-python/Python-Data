import openpyxl

#load the workbook

wb=openpyxl.load_workbook("data.xlsx")

#check the sheets availabe in file

sn=wb.sheetnames

print(sn)

#Check active sheet

print(wb.active.title)

#Read data from the sheet

# Metode 1

print(wb['Name']['A2'])

# Methode 2

sh1=wb['Name']

print(sh1['B2'].value)

# Methode 3 object.cell(row num,col num).value

print(sh1.cell(2,1).value)

# methode 4

print(sh1.cell(row=1,column=1).value)

